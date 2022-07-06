from openpyxl import Workbook
import datetime


wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
assert ws['A1'] == 42
assert isinstance(ws['A1'].value, int)

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()
